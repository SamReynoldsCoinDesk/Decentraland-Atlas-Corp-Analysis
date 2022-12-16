import json
import requests
from openpyxl import load_workbook
import openpyxl


def save_data_to_excel(data_list, filename):
    """Save data to excel"""

    workbook_name = f"{filename}.xlsx"
    try:
        wb = load_workbook(workbook_name)
    except:
        wb = openpyxl.Workbook()
    page = wb.active
    page.append(data_list)
    wb.save(filename=workbook_name)


def get_username(id):
    """Get Username from api"""

    headers = {
        "authority": "peer.decentraland.org",
        "accept": "*/*",
        "accept-language": "en-US,en;q=0.9",
        "cache-control": "no-cache",
        "origin": "https://market.decentraland.org",
        "pragma": "no-cache",
        "referer": "https://market.decentraland.org/",
        "sec-ch-ua": '"Chromium";v="106", "Google Chrome";v="106", "Not;A=Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-site",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36",
    }

    params = {
        "id": id,
    }

    response = requests.get(
        "https://peer.decentraland.org/lambdas/profiles", params=params, headers=headers
    )

    try:
        return response.json()[0]["avatars"][0]["name"]
    except:
        returned_data = id[0:6]
        print(returned_data)
        return returned_data


def check_total(id):
    """Get total nft from api"""

    headers = {
        "authority": "nft-api.decentraland.org",
        "accept": "application/json, text/plain, */*",
        "accept-language": "en-US,en;q=0.9",
        "cache-control": "no-cache",
        "origin": "https://market.decentraland.org",
        "pragma": "no-cache",
        "referer": "https://market.decentraland.org/",
        "sec-ch-ua": '"Chromium";v="106", "Google Chrome";v="106", "Not;A=Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-site",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36",
    }

    params = {
        "first": "24",
        "skip": "0",
        "sortBy": "newest",
        "owner": id,
    }
    try:
        response = requests.get(
            "https://nft-api.decentraland.org/v1/nfts", params=params, headers=headers
        )
        if response.json()["total"] >= 1:
            return "Yes"
        else:
            return "No"
    except:
        print(f"Error for {id}")
        return "No"


def read_json_file(json_filename):
    """Read json file"""
    f = open(json_filename)
    data = json.load(f)
    f.close()
    return data


def main():
    """Main function"""
    output_filename = "output"
    input_json_filename = "input_json.json"
    save_data_to_excel(
        [
            "Time on site (Mins)",
            "Wallet ID" "Username",
            "Wallet Link",
            "Has NFTs in Wallet",
        ],
        "output",
    )

    json_data = read_json_file(input_json_filename)

    for each_data in json_data:
        id = each_data["_id"]

        minutes = each_data["minutes"]

        link = f"https://market.decentraland.org/accounts/{id}"

        has_nft = check_total(id)
        username = get_username(id)
        data_list = [minutes, id, username, link, has_nft]
        print(data_list)
        save_data_to_excel(data_list, output_filename)


if __name__ == "__main__":
    main()
